from openpyxl import load_workbook


def insert_sticker(keyword, sticker_id=None, reply_text=None):
    row = stickers_page.max_row + 1
    stickers_page.cell(row=row, column=1).value = keyword
    stickers_page.cell(row=row, column=2).value = sticker_id
    stickers_page.cell(row=row, column=3).value = reply_text
    bd.save('base.xlsx')

    stickers[keyword] = sticker_id
    replies[keyword] = reply_text


bd = load_workbook('base.xlsx')
for sheet in bd:
    print(sheet.title)
stickers_page = bd['stickers']
user_page = bd['Users']

stickers = {}
replies = {}


#def insert_user(user_id, name, sex, grade):
    # '''
    # вносит нового пользователя в базу данных
    # '''
    # user_page = bd['Users']
    # row = user_page.max_row + 1
    # user_page.cell(row=row, column=1).value = user_id
    # user_page.cell(row=row, column=2).value = name
    # user_page.cell(row=row, column=3).value = sex
    # user_page.cell(row=row, column=3).value = grade
    # bd.save('base.xlsx')


def in_database(user: int) -> bool:
    '''
    возвращает True, если id пользователь есть в database
    '''
    bd = {}
    user_page = bd['Users']
    for row in range(1, user_page.max_row + 1):
        if user == user_page.cell(row=row, column=1).value:
            return True
        return False



# for row in range(1, stickers_page.max_row + 1):
#     keyword = stickers_page.cell(row=row, column=1).value
#     sticker_id = stickers_page.cell(row=row, column=2).value
#     reply_text = stickers_page.cell(row=row, column=3).value
#     stickers[keyword] = sticker_id
#     replies[keyword] = reply_text


def insert_user(*args):
    '''
    Записывает нового пользователя в bd
    :param args: принимает произвольное количество аргументов
    :return:
    '''
    user_id = args[o]
    name = args[1]
    sex = args[2]
    grade = args[3]
    row = users_page.max_row + 1
    users_page.cell(row=row, column=1).value = user_id
    users_page.cell(row=row, column=2).value = name
    users_page.cell(row=row, column=3).value = sex
    users_page.cell(row=row, column=4).value = grade
    bd.save('base.xlsx')


if __name__ == '__main__':
    main()