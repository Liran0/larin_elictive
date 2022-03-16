from telegram.ext import Updater, Filters, MessageHandler,CallbackContext
from key import TOKEN
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from openpyxl import load_workbook


def new_sticker(updade: Update, context: CallbackContext):
    sticker_id = updade.message.sticker.file_id
    for keyword in stickers:
        if sticker_id == stickers[keyword]:
            update.message.reply_text(' у меня тоже такой есть ')
            updade.message.reply_sticker(sticker_id)
            break
    else:
        context.user_data['new_sticker'] = sticker_id
        updade.message.reply_text('скажи мне ключевое слово для стикера')


def new_keyword(updade: Update, context: CallbackContext):
    if new_sticker not in context.user_data:
        sey_smth(updade, context)
    else:
        keyword = update.massage.text
        sticker_id = context.user_data['new sticker']
        insert_sticker(keyword, sticker_id)
        context.user_data.clear()


def insert_sticker(keyword, sticker_id=None, reply_text=None):
    row = stickers_page.max_row +1
    sticker_page.cell(row=row, coum=1).value = keyword
    sticker_page.cell(row=row, coum=1).value = sticker_id
    sticker_page.cell(row=row, coum=1).value = reply_text
    bd.save('base.xlsx')

    stickers[keyword] = sticker_id
    replies[keyword] = reply_text


bd = load_workbook('base.xlsx')


#for row in range(2, stickers_page.max_row +1):


def main():
    updater = Updater(
        token=TOKEN,
        use_context=True
)
    text_handler = MessageHandler(Filters.text('да'), sey_no)
    dispatcher = updater.dispatcher
    keybord_handler = MessageHandler(Filters.text('клавиатура'), keybord)
    exo_handler = MessageHandler(Filters.all, exo)
    hallo_handler = MessageHandler(Filters.text('Привет, привет'), say_hallo)
    sticker_hendler = MessageHandler(Filters.sticker,  new_sticker)

    dispatcher.add_handler(hallo_handler)
    dispatcher.add_handler(keybord_handler)
    dispatcher.add_handler(exo_handler)
    dispatcher.add_handler(text_handler)

    updater.start_polling()
    print('Бот успешно запустился')
    updater.idle()


def say_hallo(update, context):
    name = update.message.from_user.first_name
    text = update.message.text
    update.message.reply_text(text="привет")


def exo(update: Update, context: CallbackContext):
    name = update.message.from_user.first_name
    user_id = update.message.chat_id
    text = update.message.text if update.message.text else "но текста нет"
    update.message.reply_text(text=f'Привет, {name}!\n'
                              f'Твой id: {user_id}\n'
                              f'Ты написал какой-то ****: {text}\n')


def keybord(update: Update, context: CallbackContext):
    buttons = [
        ['1', '2', '3'],
        ['привет', 'да', 'пока']
    ]
    keys = ReplyKeyboardMarkup
    update.message.reply_text(
        text='смотри, есть клава ',
        reply_markup=ReplyKeyboardMarkup(
            buttons,
            resize_keyboard=True

        )
    )
def say_smth(update: Update, context: CallbackContext):
    name = update.message.from_user.first_name
    text = update.message.text
    for keyword in stickers:
        if keyword in text:
            if stickers[keyword]:
                update.message.reply_sticker(stickers[keyword])
            if replies[keyword]:
                update.message.reply_text(replies[keyword].format(name))
            break
    else:
        exo(update, context)


def sey_no(update: Update, context: CallbackContext):
    name = update.message.from_user.first_name
    update.message.reply_text(f"gjrf, {name}.")
    update.message.reply_sticker()


if __name__ == '__main__':
    main()
    print(stickers)
    insert_sticker('до свидания',reply_text='ldldldl')